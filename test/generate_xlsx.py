# -*- utf-8 -*-

from openpyxl import Workbook, load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import random
import string


def get_isbn():
    v0 = random.randint(1, 1000)
    v1 = random.randint(1, 10)
    v2 = random.randint(1, 1000)
    v3 = random.randint(1, 100000)
    v4 = random.randint(1, 10)
    return str(v0) + '-' + str(v1) + '-' + str(v2) + '-' + str(v3) + '-' + str(v4)


def get_name():
    return ''.join(random.choice(string.ascii_uppercase + string.digits +
                                 string.ascii_lowercase) for _ in range(10))


def get_price():
    return str(random.randint(1, 100))


def generate_xls_files(dirpath, count):
    for num in range(1, count):
        wb = Workbook()
        dest_filename = '{0}.xlsx'.format(num)
        dest = dirpath + dest_filename
        ws = wb.active
        ws.title = "Example"
        for row in range(1, 101):
            for col in range(1, 4):
                value = get_column_letter(col)
                if col == 1:
                    value = get_isbn()
                elif col == 2:
                    value = get_name()
                else:
                    value = get_price()
                _ = ws.cell(column=col, row=row, value="{0}".format(value))
        wb.save(filename=dest)


if __name__ == '__main__':
    generate_xls_files('E:/sw5cc/test/data/xlsx/', 10)
